from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.utils import get_column_letter

from src.scorer import is_audited
from src.utils import ensure_dir


LEAD_COLUMNS = [
    "business_name",
    "sector",
    "city",
    "rating",
    "review_count",
    "website_url",
    "decision_bucket",
    "opportunity_priority",
    "outreach_priority",
    "data_quality_status",
    "data_quality_reason",
    "normalized_business_name",
    "final_opportunity_score",
    "business_fit_score",
    "website_pain_score",
    "pain_gate_pass",
    "pain_gate_reason",
    "outreach_decision",
    "outreach_status",
    "load_confidence",
    "audit_error_type",
    "pagespeed_performance_score",
    "visual_audit_status",
    "browser_loads",
    "visual_pain_score",
    "visual_pain_reasons",
    "mobile_has_horizontal_scroll",
    "above_fold_cta_visible",
    "above_fold_phone_visible",
    "above_fold_form_visible",
    "desktop_screenshot_path",
    "mobile_screenshot_path",
    "contact_form_found",
    "cta_found",
    "text_length",
    "old_website_signals",
    "google_phone",
    "email_address",
    "google_maps_url",
    "priority",
    "country",
    "business_fit_status",
    "candidate_type",
    "website_type",
    "audit_queue",
    "final_review",
    "business_fit_reasons",
    "reject_reason",
    "prefilter_score",
    "prefilter_status",
    "prefilter_reasons",
    "reason",
    "suggested_outreach_angle",
    "opportunity_score",
    "website_weakness_score",
    "business_value_score",
    "trust_mismatch_score",
    "reachability_score",
    "website_loads",
    "uses_https",
    "meta_description",
    "homepage_title",
    "title",
    "http_status_code",
    "audit_error_message",
    "final_url",
    "email_found",
    "phone_found",
    "website_email",
    "website_phone",
    "pagespeed_performance",
    "pagespeed_seo",
    "pagespeed_accessibility",
    "pagespeed_best_practices",
    "lcp",
    "cls",
    "address",
    "place_id",
    "query",
]

RAW_COLUMNS = [
    "business_name",
    "sector",
    "city",
    "rating",
    "review_count",
    "website_url",
    "website_type",
    "decision_bucket",
    "opportunity_priority",
    "data_quality_status",
    "data_quality_reason",
    "candidate_type",
    "business_fit_score",
    "business_fit_status",
    "google_phone",
    "google_maps_url",
    "address",
    "place_id",
    "query",
    "outreach_status",
    "audit_queue",
    "reject_reason",
    "business_fit_reasons",
    "prefilter_score",
    "prefilter_status",
    "prefilter_reasons",
]


def export_leads(
    leads: list[dict[str, Any]],
    output_dir: str | Path = "output",
    total_discovered: int | None = None,
    filename_prefix: str = "leads_output",
    output_path: str | Path | None = None,
) -> Path:
    ensure_dir(output_dir)
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = Path(output_dir) / f"{filename_prefix}_{timestamp}.xlsx"
    else:
        output_path = Path(output_path)
        ensure_dir(output_path.parent)

    readme_df = _readme_dataframe()
    current_summary_df = _current_run_summary_dataframe(leads)
    send_now_df = _lead_dataframe(_current_run_bucket(leads, "send_now"), sort_by="final_opportunity_score")
    no_website_df = _lead_dataframe(_current_run_bucket(leads, "no_website_offer"), sort_by="business_fit_score")
    platform_offer_df = _lead_dataframe(_current_run_bucket(leads, "platform_offer"), sort_by="business_fit_score")
    data_quality_df = _lead_dataframe(_data_quality_review_leads(leads), sort_by="business_fit_score")
    manual_review_df = _lead_dataframe(_current_run_bucket(leads, "manual_review"), sort_by="final_opportunity_score")
    needs_browser_df = _lead_dataframe(_current_run_bucket(leads, "needs_browser_check"), sort_by="final_opportunity_score")
    visual_review_df = _lead_dataframe(_visual_review_leads(leads), sort_by="visual_pain_score")
    looks_fine_df = _lead_dataframe(_current_run_bucket(leads, "looks_fine"), sort_by="final_opportunity_score")
    hard_skip_df = _lead_dataframe(_current_run_bucket(leads, "hard_skip", require_clean=False), sort_by="business_fit_score")
    audited_run_df = _lead_dataframe(_audit_queue_leads(leads), sort_by="final_opportunity_score")
    current_raw_df = _raw_dataframe([lead for lead in leads if lead.get("current_run")])
    current_candidates_df = _lead_dataframe(_current_run_business_fit(leads), sort_by="business_fit_score")
    all_db_df = _lead_dataframe(leads, sort_by="final_opportunity_score")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        readme_df.to_excel(writer, index=False, sheet_name="README")
        current_summary_df.to_excel(writer, index=False, sheet_name="Current Run Summary")
        audited_run_df.to_excel(writer, index=False, sheet_name="Audited This Run")
        send_now_df.to_excel(writer, index=False, sheet_name="Send Now")
        no_website_df.to_excel(writer, index=False, sheet_name="No Website Offer")
        platform_offer_df.to_excel(writer, index=False, sheet_name="Platform Website Offer")
        data_quality_df.to_excel(writer, index=False, sheet_name="Data Quality Review")
        manual_review_df.to_excel(writer, index=False, sheet_name="Manual Review")
        needs_browser_df.to_excel(writer, index=False, sheet_name="Needs Browser Check")
        visual_review_df.to_excel(writer, index=False, sheet_name="Visual Review")
        looks_fine_df.to_excel(writer, index=False, sheet_name="Looks Fine")
        hard_skip_df.to_excel(writer, index=False, sheet_name="Hard Skip")
        current_raw_df.to_excel(writer, index=False, sheet_name="Current Run - Raw")
        current_candidates_df.to_excel(writer, index=False, sheet_name="Current Run - Candidates")
        all_db_df.to_excel(writer, index=False, sheet_name="All Database")

        for sheet_name in writer.book.sheetnames:
            _format_sheet(writer.book[sheet_name])

    return output_path


def _decision_bucket(lead: dict[str, Any]) -> str:
    quality_status = lead.get("data_quality_status")
    if quality_status in {"review", "noise"}:
        return "data_quality_review"

    candidate_type = lead.get("candidate_type")
    business_fit = lead.get("business_fit_status", lead.get("prefilter_status"))
    audit = lead.get("website_audit", {}) or {}
    website_type = lead.get("website_type")

    if business_fit == "skip" or candidate_type in {"weak_fit", "skip"}:
        return "hard_skip"
    if lead.get("reject_reason"):
        return "hard_skip"

    if candidate_type == "no_website_candidate" or website_type == "missing":
        return "no_website_offer"
    if candidate_type == "platform_candidate" or website_type in {"social_media", "booking_platform"}:
        return "platform_offer"

    if not is_audited(lead):
        return ""

    if audit.get("load_confidence") == "blocked_or_uncertain":
        return "needs_browser_check"

    decision = lead.get("outreach_decision")
    if decision == "send_now":
        return "send_now"
    if decision == "manual_review":
        return "manual_review"
    if decision == "looks_fine":
        return "looks_fine"
    return "manual_review"


def _opportunity_priority(lead: dict[str, Any]) -> str:
    bfs = lead.get("business_fit_score") or 0
    final = lead.get("final_opportunity_score") or 0
    score = max(bfs, final)
    if score >= 60:
        return "high"
    if score >= 48:
        return "medium"
    if score >= 30:
        return "low"
    return "skip"


def _outreach_priority(lead: dict[str, Any]) -> str:
    audit = lead.get("website_audit", {}) or {}
    candidate_type = lead.get("candidate_type")
    decision = lead.get("outreach_decision")
    load_conf = audit.get("load_confidence")
    has_phone = bool(lead.get("google_phone"))
    has_email = bool(audit.get("email_address") or audit.get("email_found"))
    has_form = bool(audit.get("contact_form_found"))
    has_contact = has_phone or has_email or has_form
    bfs = lead.get("business_fit_score") or 0

    if decision == "send_now":
        return "high"
    if candidate_type == "no_website_candidate":
        if has_phone and bfs >= 55:
            return "medium"
        return "low" if has_phone else "skip"
    if candidate_type == "platform_candidate":
        if has_contact and bfs >= 55:
            return "medium"
        return "low" if has_contact else "skip"
    if load_conf == "blocked_or_uncertain":
        return "low"
    if decision == "manual_review":
        return "medium" if has_contact else "low"
    if decision == "looks_fine":
        return "low"
    return "skip"


def _current_run_bucket(
    leads: list[dict[str, Any]],
    bucket: str,
    *,
    require_clean: bool = True,
) -> list[dict[str, Any]]:
    return [
        lead
        for lead in leads
        if lead.get("current_run")
        and _decision_bucket(lead) == bucket
        and (not require_clean or lead.get("data_quality_status") == "clean")
    ]


def _audit_queue_leads(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [lead for lead in leads if lead.get("audit_queue")]


def _data_quality_review_leads(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        lead
        for lead in leads
        if lead.get("current_run")
        and lead.get("data_quality_status") in {"review", "noise"}
    ]


def _visual_review_leads(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    selected: list[dict[str, Any]] = []
    for lead in leads:
        if not lead.get("current_run"):
            continue
        if lead.get("data_quality_status") != "clean":
            continue
        visual = lead.get("visual_audit", {}) or {}
        if visual.get("visual_audit_status") != "audited":
            continue
        visual_pain = int(visual.get("visual_pain_score") or 0)
        if visual_pain >= 20:
            selected.append(lead)
            continue
        if lead.get("outreach_decision") == "looks_fine" and visual_pain > 0:
            selected.append(lead)
    return selected


def _current_run_business_fit(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        lead
        for lead in leads
        if lead.get("current_run")
        and lead.get("business_fit_status", lead.get("prefilter_status")) != "skip"
    ]


def _lead_dataframe(
    leads: list[dict[str, Any]],
    sort_by: str = "final_opportunity_score",
) -> pd.DataFrame:
    rows = [_flatten_lead(lead) for lead in leads]
    df = pd.DataFrame(rows).reindex(columns=LEAD_COLUMNS)
    if not df.empty:
        primary = sort_by if sort_by in df.columns else "business_fit_score"
        df = df.sort_values(
            by=[primary, "business_fit_score"],
            ascending=[False, False],
            na_position="last",
        )
    return df


def _raw_dataframe(leads: list[dict[str, Any]]) -> pd.DataFrame:
    rows = [_flatten_lead(lead) for lead in leads]
    df = pd.DataFrame(rows).reindex(columns=RAW_COLUMNS)
    if not df.empty:
        df = df.sort_values("business_fit_score", ascending=False, na_position="last")
    return df


def audited_leads(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [lead for lead in leads if is_audited(lead)]


def send_candidate_leads(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        lead
        for lead in leads
        if is_audited(lead)
        and lead.get("outreach_decision") == "send_now"
    ]


def manual_review_leads(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        lead
        for lead in leads
        if is_audited(lead)
        and lead.get("outreach_decision") == "manual_review"
    ]


def looks_fine_leads(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        lead
        for lead in leads
        if is_audited(lead)
        and lead.get("outreach_decision") == "looks_fine"
    ]


def needs_browser_check_leads(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        lead
        for lead in leads
        if is_audited(lead)
        and (lead.get("website_audit", {}) or {}).get("load_confidence") == "blocked_or_uncertain"
    ]


def no_website_leads(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    selected: list[dict[str, Any]] = []
    for lead in leads:
        candidate_type = lead.get("candidate_type")
        if candidate_type in {"weak_fit", "skip"}:
            continue
        status = lead.get("business_fit_status", lead.get("prefilter_status"))
        if status == "skip":
            continue
        if candidate_type == "no_website_candidate" or lead.get("website_type") == "missing":
            selected.append(lead)
    return selected


def platform_leads(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    selected: list[dict[str, Any]] = []
    for lead in leads:
        candidate_type = lead.get("candidate_type")
        if candidate_type in {"weak_fit", "skip"}:
            continue
        status = lead.get("business_fit_status", lead.get("prefilter_status"))
        if status == "skip":
            continue
        if candidate_type == "platform_candidate" or lead.get("website_type") in {
            "social_media",
            "booking_platform",
        }:
            selected.append(lead)
    return selected


def _flatten_lead(lead: dict[str, Any]) -> dict[str, Any]:
    audit = lead.get("website_audit", {}) or {}
    pagespeed = lead.get("pagespeed", {}) or {}
    visual = lead.get("visual_audit", {}) or {}
    return {
        "business_name": lead.get("business_name"),
        "sector": lead.get("sector"),
        "city": lead.get("city"),
        "country": lead.get("country"),
        "rating": lead.get("rating"),
        "review_count": lead.get("review_count"),
        "website_url": lead.get("website_url"),
        "website_type": lead.get("website_type"),
        "decision_bucket": _decision_bucket(lead),
        "opportunity_priority": _opportunity_priority(lead),
        "outreach_priority": _outreach_priority(lead),
        "data_quality_status": lead.get("data_quality_status"),
        "data_quality_reason": lead.get("data_quality_reason"),
        "normalized_business_name": lead.get("normalized_business_name"),
        "final_opportunity_score": lead.get("final_opportunity_score"),
        "business_fit_score": lead.get("business_fit_score"),
        "website_pain_score": lead.get("website_pain_score"),
        "pain_gate_pass": lead.get("pain_gate_pass"),
        "pain_gate_reason": lead.get("pain_gate_reason"),
        "outreach_decision": lead.get("outreach_decision"),
        "outreach_status": lead.get("outreach_status"),
        "load_confidence": audit.get("load_confidence"),
        "audit_error_type": audit.get("audit_error_type"),
        "pagespeed_performance_score": pagespeed.get("performance_score"),
        "visual_audit_status": visual.get("visual_audit_status"),
        "browser_loads": visual.get("browser_loads"),
        "visual_pain_score": visual.get("visual_pain_score"),
        "visual_pain_reasons": _format_signals(visual.get("visual_pain_reasons", [])),
        "mobile_has_horizontal_scroll": visual.get("mobile_has_horizontal_scroll"),
        "above_fold_cta_visible": visual.get("above_fold_cta_visible"),
        "above_fold_phone_visible": visual.get("above_fold_phone_visible"),
        "above_fold_form_visible": visual.get("above_fold_form_visible"),
        "desktop_screenshot_path": visual.get("desktop_screenshot_path"),
        "mobile_screenshot_path": visual.get("mobile_screenshot_path"),
        "contact_form_found": audit.get("contact_form_found"),
        "cta_found": audit.get("cta_found"),
        "text_length": audit.get("text_length"),
        "old_website_signals": _format_signals(audit.get("old_website_signals", [])),
        "google_phone": lead.get("google_phone"),
        "email_address": audit.get("email_address"),
        "google_maps_url": lead.get("google_maps_url"),
        "priority": lead.get("priority"),
        "business_fit_status": lead.get("business_fit_status"),
        "candidate_type": lead.get("candidate_type"),
        "audit_queue": lead.get("audit_queue"),
        "final_review": lead.get("final_review"),
        "business_fit_reasons": lead.get("business_fit_reasons"),
        "reject_reason": lead.get("reject_reason"),
        "prefilter_score": lead.get("prefilter_score"),
        "prefilter_status": lead.get("prefilter_status"),
        "prefilter_reasons": lead.get("prefilter_reasons"),
        "reason": lead.get("reason"),
        "suggested_outreach_angle": lead.get("suggested_outreach_angle"),
        "opportunity_score": lead.get("opportunity_score"),
        "website_weakness_score": lead.get("website_weakness_score"),
        "business_value_score": lead.get("business_value_score"),
        "trust_mismatch_score": lead.get("trust_mismatch_score"),
        "reachability_score": lead.get("reachability_score"),
        "website_loads": audit.get("website_loads"),
        "uses_https": audit.get("uses_https"),
        "meta_description": audit.get("meta_description"),
        "homepage_title": audit.get("homepage_title"),
        "title": audit.get("title"),
        "http_status_code": audit.get("http_status_code"),
        "audit_error_message": audit.get("audit_error_message"),
        "final_url": audit.get("final_url"),
        "email_found": audit.get("email_found"),
        "phone_found": audit.get("phone_found"),
        "website_email": audit.get("email_address"),
        "website_phone": audit.get("phone_text"),
        "pagespeed_performance": pagespeed.get("performance_score"),
        "pagespeed_seo": pagespeed.get("seo_score"),
        "pagespeed_accessibility": pagespeed.get("accessibility_score"),
        "pagespeed_best_practices": pagespeed.get("best_practices_score"),
        "lcp": pagespeed.get("lcp"),
        "cls": pagespeed.get("cls"),
        "address": lead.get("address"),
        "place_id": lead.get("place_id"),
        "query": lead.get("query"),
    }


def _readme_dataframe() -> pd.DataFrame:
    rows = [
        {"section": "Purpose", "text": "This workbook is a sales decision screen, not a database dump. Top sheets are scoped to the current pipeline run."},
        {"section": "Sheet order", "text": "Current-run sheets come first. Database-wide sheets are at the end."},
        {"section": "Send Now", "text": "Audited custom websites with clear pain. The only sheet meant for immediate outreach without manual review."},
        {"section": "No Website Offer", "text": "Leads with no website at all. A separate opportunity type, not a skip. Reach out by Google phone where available."},
        {"section": "Platform Website Offer", "text": "Leads using social media or a booking platform instead of a real website. A separate opportunity type."},
        {"section": "Data Quality Review", "text": "Current-run leads flagged as 'review' or 'noise' by the data-quality filter. 'noise' = unusable record (e.g. business_name == city). 'review' = usually keyword-stuffed Google Places SEO names that may still be a real business but need a human eye. Visual audit only runs on 'clean' leads."},
        {"section": "Manual Review", "text": "Audited custom websites with ambiguous signals. Open the site and decide manually."},
        {"section": "Needs Browser Check", "text": "The auditor could NOT confirm the site is dead. It may be blocked, behind SSL/redirect issues, timing out, or rate-limited. Always open in a browser before classifying."},
        {"section": "Visual Review", "text": "Sites that loaded fine on HTTP but the headless-browser visual audit found possible visible design pain (no above-fold CTA, mobile horizontal scroll, very thin above-fold text, missing phone/form). Open the screenshots before deciding."},
        {"section": "Looks Fine", "text": "Audited custom websites with no clear pain detected. Lower outreach priority."},
        {"section": "Hard Skip", "text": "Rejected by business-fit rules, weak fit, or out of scope."},
        {"section": "Audited This Run", "text": "The exact leads HTTP-audited in this pipeline run."},
        {"section": "Current Run - Raw", "text": "Every lead discovered in this pipeline run."},
        {"section": "Current Run - Candidates", "text": "Current-run leads that passed business-fit prefiltering."},
        {"section": "All Database", "text": "Every lead in SQLite, sorted by final_opportunity_score."},
        {"section": "decision_bucket", "text": "Derived. Values: send_now, no_website_offer, platform_offer, manual_review, needs_browser_check, looks_fine, hard_skip, or empty (pending audit)."},
        {"section": "opportunity_priority", "text": "Derived. How interesting the business is as a sales target. high / medium / low / skip."},
        {"section": "outreach_priority", "text": "Derived. Whether the lead is contact/action ready right now. Only 'high' for audited custom websites with clear pain."},
    ]
    return pd.DataFrame(rows)


def _current_run_summary_dataframe(leads: list[dict[str, Any]]) -> pd.DataFrame:
    current = [lead for lead in leads if lead.get("current_run")]
    bucket_counts = Counter(_decision_bucket(lead) for lead in current)
    quality_counts = Counter(lead.get("data_quality_status") for lead in current)
    audit_run = [lead for lead in leads if lead.get("audit_queue")]

    rows: list[dict[str, Any]] = [
        {"metric": "current_run_total", "value": len(current)},
        {"metric": "audited_this_run", "value": len(audit_run)},
        {"metric": "data_quality_clean (current run)", "value": quality_counts.get("clean", 0)},
        {"metric": "data_quality_review (current run)", "value": quality_counts.get("review", 0)},
        {"metric": "data_quality_noise (current run)", "value": quality_counts.get("noise", 0)},
        {"metric": "send_now (current run)", "value": bucket_counts.get("send_now", 0)},
        {"metric": "no_website_offer (current run)", "value": bucket_counts.get("no_website_offer", 0)},
        {"metric": "platform_offer (current run)", "value": bucket_counts.get("platform_offer", 0)},
        {"metric": "manual_review (current run)", "value": bucket_counts.get("manual_review", 0)},
        {"metric": "needs_browser_check (current run)", "value": bucket_counts.get("needs_browser_check", 0)},
        {"metric": "visual_review (current run)", "value": len(_visual_review_leads(leads))},
        {"metric": "looks_fine (current run)", "value": bucket_counts.get("looks_fine", 0)},
        {"metric": "hard_skip (current run)", "value": bucket_counts.get("hard_skip", 0)},
        {"metric": "pending_audit (current run)", "value": bucket_counts.get("", 0)},
        {"metric": "total_db_leads", "value": len(leads)},
    ]

    rows.extend(
        {"metric": f"current run - sector: {sector}", "value": count}
        for sector, count in Counter(lead.get("sector", "Unknown") for lead in current).most_common(10)
    )
    rows.extend(
        {"metric": f"current run - city: {city}", "value": count}
        for city, count in Counter(lead.get("city", "Unknown") for lead in current).most_common(10)
    )

    return pd.DataFrame(rows)


def _format_signals(value: Any) -> str:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return str(value or "")


def _format_sheet(worksheet: Any) -> None:
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 60))
        worksheet.column_dimensions[column_letter].width = max(max_length + 2, 12)
